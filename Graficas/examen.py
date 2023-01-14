import csv
import matplotlib.pyplot as plt
import numpy as np
from collections import Counter
import statistics
from openpyxl import Workbook, load_workbook


# ----------------------------------------------------------------
# Nombre del programa: Examen Parcial 2
# Objetivo / propósito: Crea un programa en Python, que permita analizar el dataset de registros de las ventas de un e-commerce
# Parámetros de entrada: Dataset de registros de ventas
# Parámetros de salida: Lo que pide el examen xd
# Lenguaje: Python
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------

tienda = {}

# ----------------------------------------------------------------
# Nombre de la función: Menu
# Objetivo / propósito: Poner un menu de opciones para que el usuario pueda elegir la opcion
# Parámetros de entrada: N/A
# Parámetros de salida: N/A
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------
def menu():
    print("Bienvenido al menu de tiendas de supermecados, ¿que te gustaria hacer?")
    print("--- [Menu] ---")
    print("[1] Importar datos")
    print("[2] Numero de pedidos por paises")
    print("[3] Minimo, maximo, media, desviacion estandar de UnitPrice")
    print("[4] Calcula el precio total de los primeros 10 productos")
    print("[5] Cual es el cliente con mas compras?")
    print("[6] Que tipo de sistema es?")
    print("[0] Salir")
menu()
    
# ----------------------------------------------------------------
# IMPORTANTE!!!! DESCOMENTAR ESTE COMENTARIO PARA PODER TRANSFORMAR EL DATA SET CSV A XLSX, MUY IMPORTANTE SINO EL CODIGO NO PODRA LEER EL DATASET
#----------------------------------------------------------------    
    
'''import openpyxl
# ----------------------------------------------------------------
# Se convierte un archivo csv a un xlsx en caso de que los datos esten en un csv
#----------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active

with open('dataset_examen2.csv', encoding='latin-1') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

wb.save('dataset_examen2.xlsx')'''


# ----------------------------------------------------------------
# Nombre de la función: Importar datos
# Objetivo / propósito: Importar datos de un archivo xlsx a un diccionario
# Parámetros de entrada: Archivo xlsx
# Parámetros de salida: Diccionario tienda
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------
def opc1():
    wb = load_workbook('dataset_examen2.xlsx')
    ws = wb.active
    maxi = ws.max_row
    rows = ws.iter_rows(min_row=2, max_row = maxi, min_col = 1, max_col= 8)
    for InvoiceNo, StockCode, Description, Quantity, InvoiceDate, UnitPrice, customerID, Country in rows:
        i = (len(tienda)+1)
        tienda[i] = {}
        tienda[i]['InvoiceNo'] = InvoiceNo.value
        tienda[i]['StockCode'] = StockCode.value
        tienda[i]['Description'] = Description.value
        tienda[i]['Quantity'] = Quantity.value
        tienda[i]['InvoiceDate'] = InvoiceDate.value
        tienda[i]['UnitPrice'] = UnitPrice.value
        tienda[i]['CustomerID'] = customerID.value
        tienda[i]['Country'] = Country.value
    print("¡Se han importado los datos de la tienda!")

# ----------------------------------------------------------------
# Nombre de la función: PROBLEMA 1
# Objetivo / propósito: Numero de pedidos por paises
# Parámetros de entrada: Diccionario con todos los datos del dataset
# Parámetros de salida: El contador de todos los paises con su numero de pedidos 
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------

#COMPLETADO PROBLEMA 1
def opc2():
    paises = []
    for i in range(len(tienda)):
        paises.append(tienda[i+1]['Country'])
    pedido = dict(Counter(paises))
    print('La cantidad de pedidos por paises es el siguiente:', pedido)

# ----------------------------------------------------------------
# Nombre de la función: PROBLEMA 2
# Objetivo / propósito: Minimo, maximo, media, desviacion estandar de UnitPrice
# Parámetros de entrada: Diccionario con todos los datos del dataset
# Parámetros de salida: Minimo, maximo, media y desviacion estandar de UnitPrice
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------
#COMPLETADO PROBLEMA 2
def opc3():
    x1 = []
    y1 = []
    y2 = []
    y3 = []
    a = 0
    for i in range(len(tienda)):
        x1.append(tienda[i+1]['UnitPrice'])
    y1 = x1
    y2 = x1
    y1.sort(key=float)
    maximo = y1.pop()
    print("El maximo valor es: ", maximo)
    y2.reverse()
    minimo = y2.pop()
    print("El minimo valor es: ", minimo)
    for i in range(len(tienda)):
        x = float(tienda[i+1]['UnitPrice'])
        a = a + x
    total = a / (len(tienda))
    print("La media de UnitPrice es:",total)
    for item in y1:
        y3.append(float(item))
    std = statistics.pstdev((y3))
    print("La desviación estandar es:", std)
    
    
# ----------------------------------------------------------------
# Nombre de la función: PROBLEMA 3
# Objetivo / propósito: Calcula el precio total de los primeros 10 productos
# Parámetros de entrada: Diccionario con todos los datos del dataset
# Parámetros de salida: Precio total de los 10 productos
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------
#COMPLETADO PROBLEMA 3
def opc4():
    i = 1
    precio = 0
    preciofinal = 0
    for i in range (10):
        print(f"Description:{tienda[i+1]['Description']}, Quantity: {tienda[i+1]['Quantity']}, CustomerID: {tienda[i+1]['CustomerID']}, UnitPrice:  {tienda[i+1]['UnitPrice']}")
        c = float(tienda[i+1]['Quantity'])
        p = float(tienda[i+1]['UnitPrice'])
        precio = c * p
        preciofinal = preciofinal + precio
    print("El precio total de los primeros 10 productos es:", preciofinal)



# ----------------------------------------------------------------
# Nombre de la función: PROBLEMA 4
# Objetivo / propósito: Cliente con mas compras, su gasto y primer y ultima compra
# Parámetros de entrada: Diccionario con todos los datos del dataset
# Parámetros de salida: Cliente con mas compras, su gasto y primer y ultima compra
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------
#COMPLETADO PROBLEMA 4
def opc5():
    x1 = []
    y1 = []
    preciofinal = 0
    precio = 0
    u = 0
    tiempo = []
    for i in range(len(tienda)):
        x1.append(tienda[i+1]['CustomerID'])
    y1 = x1 
    c = dict(Counter(y1))
    b = sorted(c, key=c.get, reverse=True)
    h = b[1]
    print("La ID del cliente con mas compras es:", h)
    for i in range(len(tienda)):
        if tienda[i+1]['CustomerID'] == h:
            d = float(tienda[i+1]['Quantity'])
            p = float(tienda[i+1]['UnitPrice'])
            tiempo.append(tienda[i+1]['InvoiceDate'])
            precio = d * p
            preciofinal = preciofinal + precio
    primero = tiempo.pop(0)
    ultimo = tiempo.pop()
    print("El usuario", h, "gasto: ", preciofinal)
    print("Su primera compra fue el:", primero)
    print("Su ultima compra fue el:", ultimo)
                        

# ----------------------------------------------------------------
# Nombre de la función: PROBLEMA 5
# Objetivo / propósito: Tipo de sistema de informacion
# Parámetros de entrada: N/A
# Parámetros de salida: N/A
# Desarrolló: David Bojalil Abiti, 24-Octubre-2022
#----------------------------------------------------------------
#COMPLETADO PROBLEMA 5
def opc6():
    print("El tipo de sistema de informacion es un sistema BI (Sistema de Inteligencia Empresarial) ")
    print("Esto se debe a que se estan procesando los datos del dataset, como por ejemplo al buscar la tendencia de cuales son los paises que mas piden lo cual ")
    print("ayuda a las empresas a tomar decisiones sobre a que mercado, o en este caso pais, se deben de expander. ")
    print("En adicion, este sistema sirve para determinar la frecuencia de los precios por unidad, como lo son lo minimo, maximo, media y desviacion estandar ")
    print("Asi como tambien ser posible calcular el tiempo entre primera y ultima compra de los usuarios para hacer un estimado del tiempo minimo que un usuario tiene para comprar")


opc = int(input("Seleccione alguna opción: "))
while opc != 0:
    if (opc == 1):
        opc1()
    elif(opc == 2):
        opc2()
    elif(opc == 3):
        opc3()
    elif(opc == 4):
        opc4()
    elif(opc == 5):
        opc5()
    elif(opc == 6):
        opc6()
    else:
        print("Opción invalida")
    print()
    menu()
    opc = int(input("Seleccione alguna opción: "))
print("¡Adios! ")