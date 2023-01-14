import csv
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

# ----------------------------------------------------------------
# Nombre del programa: programa de suministros
# Objetivo / propósito: registrar, mostrar, actualizar, frecuencia, encontrar total y buscar datos
# Parámetros de entrada: diccionario de suministros
# Parámetros de salida: registrar, mostrar o buscar
# Lenguaje: Python
# Desarrolló: David Bojalil Abiti, 26-Septiembre-2022
#----------------------------------------------------------------

tienda = {}

# ----------------------------------------------------------------
# Nombre de la función: Menu
# Objetivo / propósito: Poner un menu de opciones para que el usuario pueda elegir la opcion
# Parámetros de entrada: N/A
# Parámetros de salida: N/A
# Desarrolló: David Bojalil Abiti, 26-Septiembre-2022
#----------------------------------------------------------------
def menu():
    print("Bienvenido al menu de tiendas de supermecados, ¿que te gustaria hacer?")
    print("--- [Menu] ---")
    print("[1] Importar datos de csv")
    print("[2] Mostrar datos")
    print("[3] Top 10 tiendas mas visitadas")
    print("[4] Top 5 mejores ingresos")
    print("[5] Promedio de ventas")
    print("[6] Gasto promedio por cliente")
    print("[7] Graficar")
    print("[0] Salir")
menu()

# ----------------------------------------------------------------
# Nombre de la función: Menu promedio
# Objetivo / propósito: Poner un menu de opciones para la funcion de modificar 
# Parámetros de entrada: N/A
# Parámetros de salida: N/A
# Desarrolló: David Bojalil Abiti, 28-Septiembre-2022
#----------------------------------------------------------------
def menumpromedio():
    print(" ")
    print("--- [Menu promedio] ---")
    print("[1] Gasto promedio x cliente de cierta tienda")
    print("[2] Gasto promedio x cliente de todas las tiendas")
    print("[0] Salir")
    


import openpyxl
# ----------------------------------------------------------------
# Se convierte un archivo csv a un xlsx en caso de que los datos esten en un csv
#----------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active

with open('Stores.csv') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

wb.save('stores.xlsx')

# ----------------------------------------------------------------
# Nombre de la función: Importar datos
# Objetivo / propósito: Importar datos de un archivo xlsx a un diccionario
# Parámetros de entrada: Archivo xlsx
# Parámetros de salida: Diccionario tienda
# Desarrolló: David Bojalil Abiti, 17-Octubre-2022
#----------------------------------------------------------------
def opc1():
    wb = load_workbook('stores.xlsx')
    ws = wb.active
    rows = ws.iter_rows(min_row=2, max_row = 897, min_col = 2, max_col= 5)
    for area, items, customer, sales in rows:
        i = (len(tienda)+1)
        tienda[i] = {}
        tienda[i]['area'] = area.value
        tienda[i]['items'] = items.value
        tienda[i]['customer'] = customer.value
        tienda[i]['sales'] = sales.value
    print("¡Se han importado los datos de la tienda!")

# ----------------------------------------------------------------
# Nombre de la función: Mostrar tienda
# Objetivo / propósito: Mostrar todas las tiendas
# Parámetros de entrada: diccionario de tiendas
# Parámetros de salida: todos los datos de la tienda
# Desarrolló: David Bojalil Abiti, 17-Octubre-2022
#----------------------------------------------------------------
def opc2():
    i = 1
    for i in range (len(tienda)):
        print(f"Area: {tienda[i+1]['area']}, Items available: {tienda[i+1]['items']}, Daily customers: {tienda[i+1]['customer']}, Daily sales {tienda[i+1]['sales']}")

# ----------------------------------------------------------------
# Nombre de la función: Top 10 tiendas con mas clientes
# Objetivo / propósito: Obtener el top 10 de tiendas con mas clientes
# Parámetros de entrada: Diccionario de tiendas 
# Parámetros de salida: El top 10 de tiendas con mas clientes y una grafica enseñando esas tiendas
# Desarrolló: David Bojalil Abiti, 19-Octubre-2022
#---------------------------------------------------------------- 
def opc3():
    y = []
    x1 = []
    y1 = []
    cantidad = []
    Top_valor = []
    for i in range(len(tienda)):
        x1.append(tienda[i+1]['customer'])
    y1 = x1
    y1.sort(key=int)
    print(y1)
    for c in range (0,10):
        x = y1.pop()
        cantidad.append(x)
        print(x)
        for i in range(len(tienda)):
            if (tienda[i+1]['customer'] == x):
                print(i+1)
                Top_valor.append(i+1)
    print("La ID de los top 10 tiendas mas visitadas es:", Top_valor) #Nota: Agregar i + 2 para buscar la celda que tiene ese ID :)
    '''nombre = 'Top 10 tiendas mas visitadas'
    xl = 'ID de Tiendas'
    yl = 'Cantidad de visitas'
    graficar(Top_valor,cantidad,nombre, xl, yl)'''
    

# ----------------------------------------------------------------
# Nombre de la función: Top 5 tiendas con mas ventas
# Objetivo / propósito: Obtener el top 5 de tiendas con mas ventas
# Parámetros de entrada: Diccionario de tiendas 
# Parámetros de salida: El top 5 de tiendas con mas ventas y una grafica enseñando esas tiendas
# Desarrolló: David Bojalil Abiti, 19-Octubre-2022
#---------------------------------------------------------------- 
def opc4():
    x1 = []
    y1 = []
    cantidad = []
    Top_sales = []
    for i in range(len(tienda)):
        x1.append(tienda[i+1]['sales'])
    y1 = x1
    y1.sort(key=int)
    print(y1)
    for c in range (0,5):
        x = y1.pop()
        cantidad.append(x)
        print(x)
        for i in range(len(tienda)):
            if (tienda[i+1]['sales'] == x):
                print(i+1)
                Top_sales.append(i+1)
    print("La ID de los top 5 tiendas con mas ventas es:", Top_sales) #Nota: Agregar i + 2 para buscar la celda que tiene ese ID :)
    print(cantidad)
    nombre = 'Top 5 tiendas con mas ventas'
    xl = 'ID de Tiendas'
    yl = 'Cantidad de ventas'
    graficar(Top_sales,cantidad,nombre, xl, yl)



# ----------------------------------------------------------------
# Nombre de la función: El promedio de ventas entre todas las tiendas
# Objetivo / propósito: Obtener el promedio de ventas de todas las tiendas juntas
# Parámetros de entrada: Diccionario de tiendas 
# Parámetros de salida: Promedio
# Desarrolló: David Bojalil Abiti, 19-Octubre-2022
#----------------------------------------------------------------   
def opc5():
    x = 0
    a = 0
    for i in range(len(tienda)):
        x = float(tienda[i+1]['sales'])
        a = a + x
    total = a / (len(tienda))
    
    print("El promedio de ventas es:",total)

# ----------------------------------------------------------------
# Nombre de la función: Gasto promedio x cliente
# Objetivo / propósito: Obtener el gasto promedio x cliente de 1. Una tienda en especifico o 2. Todas
# Parámetros de entrada: Diccionario de tiendas y posicion
# Parámetros de salida: Gasto promedio x cliente
# Desarrolló: David Bojalil Abiti, 19-Octubre-2022
#----------------------------------------------------------------   
def opc6():
    menumpromedio()
    opc = int(input("Seleccione alguna opción: "))
    while opc != 0:
        if (opc == 1):
            d = buscar()
            if (d != 0):
                gasto = int(tienda[d]['sales'])
                cliente = int(tienda[d]['customer'])
                total = gasto / cliente
                print("El gasto promedio x cliente de la tienda", d, "es:", total)
            else:
                print("Error, no existe esta tienda")
        elif (opc == 2):
            gastototal = 0
            clientestotales = 0
            for i in range(len(tienda)):
                gasto = int(tienda[i+1]['sales'])
                gastototal = gastototal + gasto
                cliente = int(tienda[i+1]['customer'])
                clientestotales = clientestotales + cliente
                total = gastototal / clientestotales
            print("El gasto promedio x cliente de todas las tiendas es de:",total)
        else:
            print("Opción invalida")
        menumpromedio()
        opc = int(input("Seleccione alguna opción: "))

# ----------------------------------------------------------------
# Nombre de la función: Buscar tienda
# Objetivo / propósito: Buscar tienda en diccionario
# Parámetros de entrada: diccionario de tiendas
# Parámetros de salida: Todos los datos de la tienda
# Desarrolló: David Bojalil Abiti, 19-Octubre-2022
#----------------------------------------------------------------      
def buscar():
    i = int(input("ID de tienda que quieres buscar: "))
    if (i <= len(tienda)) and (i != 0) :
        print(f"Area: {tienda[i]['area']}, Items available: {tienda[i]['items']}, Daily customers: {tienda[i]['customer']}, Daily sales {tienda[i]['sales']}")
        return i
    else:
        return 0

# ----------------------------------------------------------------
# Nombre de la función: graficar
# Objetivo / propósito: Grafciar datos
# Parámetros de entrada: datos en x y en y, nombre, xlabel y ylabel
# Parámetros de salida: Grafica con los datos
# Desarrolló: David Bojalil Abiti, 19-Octubre-2022
#----------------------------------------------------------------
def graficar(x,y,n, xl, yl):
    colores = ["#581845" , "#900C3F" , "#C70039", "#FF5733", "#FFC300"]
    plt.bar([str(s) for s in x], [int(d) for d in y] , color = colores, edgecolor = 'black')
    plt.legend()
    plt.title(n, fontweight="bold", fontname = "Times New Roman")
    plt.xticks(rotation = 30)
    plt.xlabel(xl)
    plt.ylabel(yl)
    plt.show()


opc = int(input("Seleccione alguna opción: "))
while opc != 0:
    if (opc == 1):
        opc1()
    elif(opc == 2):
        opc2()
    elif(opc == 3):
        opc3()
    elif(opc == 4):
        opc4()
    elif(opc == 5):
        opc5()
    elif(opc == 6):
        opc6()
    else:
        print("Opción invalida")
    print()
    menu()
    opc = int(input("Seleccione alguna opción: "))
print("¡Adios! ")